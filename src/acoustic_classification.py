"""
@author: Kashaf Gulzar

Main file for feature performance and subgroup wise performance metrics
"""
# -------------- Import Modules ---------------
import pandas as pd
from Classifiers import mlp, random_forest, svm
from Check_biaseness import BiasnessData, BiasnessTable
from acoustic_data_loader import Data
from Counter import ClassifierCounter, Y_testCounter
import os
from sklearn.preprocessing import MinMaxScaler
from metrics import calculate_metrics_summary
import numpy as np
from openpyxl import load_workbook

# -------------- Initilaizations ---------------
# Mode of classification
# 'Standard': Training and testing by splitting dataset into two parts
# 'Testing_Remaining': Training on AD and gender balanced data and testing on the remaining data
classification_mode = 'Standard'
# Get the directory of the current script
current_directory = os.path.dirname(os.path.realpath(__file__))
parent_directory = os.path.dirname(current_directory)

# Construct the path to the .pkl file in the parent directory
embedding_path = os.path.join(parent_directory, "Data", "Embeddings", "acoustic", "imbalanced_data_embedding.pkl")
# Embedding_name: data.x where x= MFCC, egemaps, Wav2vec_latent, Wav2vec_Hidden_y(y=1:13)
embedding_name = "Wav2vec_Hidden_13"
# labels: Ad_status or Depression_status
label_name = "Ad_status"
results_directory = os.path.join(parent_directory, "Results_final/acoustics/imbalanced_data/Embeddings_performance/"
                                                   "AD_status/wav2vec_hidden_layer_13")
# Final results path
existing_file = os.path.join(parent_directory, "Results_final/acoustics/imbalanced_data/Embeddings_performance/AD_status/embeddings_performance_summary.xlsx")

if classification_mode == 'Testing_Remaining':
    test_embedding_path = os.path.join(parent_directory, "Data", "Embeddings", "acoustic", "imbalanced_data_embedding.pkl")
    scaler = MinMaxScaler()

data_obj = Data()
random_seeds = [0, 50, 100, 150, 200]
svm = svm.SVM(x_train=None, y_train=None, x_test=None, y_test=None, save_path_opt=None, save_model_path=None)
mlp = mlp.MLP(x_train=None, y_train=None, x_test=None, y_test=None, save_path=None, save_model_path=None)
rf = random_forest.RF(x_train=None, y_train=None, x_test=None, y_test=None, save_path=None, save_model_path=None)
svm_counter = ClassifierCounter()
mlp_counter = ClassifierCounter()
rf_counter = ClassifierCounter()
ytest_counter = Y_testCounter()
svm_metrics_dict = {}
mlp_metrics_dict = {}
rf_metrics_dict = {}

results_table_path = os.path.join(results_directory, "Biasness check.xlsx")
metrics_table_path = os.path.join(results_directory, "Classifier metrics.xlsx")
# Create a Pandas Excel writer using XlsxWriter as the engine
excel_writer = pd.ExcelWriter(metrics_table_path, engine='xlsxwriter')
# --------------- main function ----------------


def main():
    data = data_obj.import_data(embedding_path)
    print(data)
    if label_name == "Depression_status":
        data = data[data['Ad_status'] == 1]
    feature_data = data_obj.data_preparation(embedding_name, data[embedding_name], data)
    labels = data[label_name]
    if classification_mode == 'Testing_Remaining':
        test_data = data_obj.import_data(test_embedding_path)  # contains files of train data as well

    for idx, seed in enumerate(random_seeds):
        if classification_mode == 'Standard':
            # test train split based on seeds
            # mode: 'Standard', 'Test_depression', 'Test_Ad'
            X_train, X_test, y_train, y_test = data_obj.preprocessing(data, feature_data, labels, seed, mode='Standard')
        elif classification_mode == 'Testing_Remaining':
            X_train, y_train = data_obj.preprocessing_different_data(feature_data, labels, seed)
            training_filenames = data.Filename.iloc[y_train.index]
            # Check if test_data is a DataFrame before accessing .index
            if isinstance(test_data, pd.DataFrame):
                pure_test_data_indices = test_data.index[~test_data['Filename'].isin(training_filenames)].tolist()
                test_data_updated = test_data.iloc[pure_test_data_indices]  # test data without the files of train data
                X_test = data_obj.data_preparation("Latent layer", test_data_updated.Wav2vec_latent, test_data)
                y_test = test_data_updated.Ad_status
                X_train = scaler.fit_transform(X_train)
                X_test = scaler.transform(X_test)
            else:
                # Handle the case where test_data is not a DataFrame as needed
                print("test_data is not a pandas DataFrame.")

        # classification
        svm.x_train, svm.y_train, svm.x_test, svm.y_test = X_train, y_train, X_test, y_test
        svm.save_path_opt = os.path.join(results_directory, "svm_result seed_{}.svg".format(seed))
        svm.model_path = os.path.join(results_directory, "svm_model seed_{}.sav".format(seed))
        svm_metrics = svm.svm_opt()
        mlp.x_train, mlp.y_train, mlp.x_test, mlp.y_test = X_train, y_train, X_test, y_test
        mlp.save_path = os.path.join(results_directory, "mlp_result seed_{}.svg".format(seed))
        mlp.model_path = os.path.join(results_directory, "mlp_model seed_{}.sav".format(seed))
        mlp_metrics = mlp.mlp_model()
        rf.x_train, rf.y_train, rf.x_test, rf.y_test = X_train, y_train, X_test, y_test
        rf.save_path = os.path.join(results_directory, "rf_result seed_{}.svg".format(seed))
        rf.model_path = os.path.join(results_directory, "rf_model seed_{}.sav".format(seed))
        rf_metrics = rf.rf_model()
        svm_metrics_dict[seed] = svm_metrics
        mlp_metrics_dict[seed] = mlp_metrics
        rf_metrics_dict[seed] = rf_metrics

        # save train and test data
        # NOTE: In case of classification_mode == 'Testing_Remaining' the indices of training data are from first dataset and
        # the indices of testing data are from the test dataset
        train_col = {'Index': []}
        train_df = pd.DataFrame(train_col)
        test_col = {'Index': []}
        test_df = pd.DataFrame(test_col)
        for idx, label in y_train.items():
            train_data = {'Index': idx}
            train_df = train_df.append(train_data, ignore_index=True)
        for idx, label in y_test.items():
            test_data = {'Index': idx}
            test_df = test_df.append(test_data, ignore_index=True)

        train_df_path = os.path.join(results_directory, "Training_data seed_{}.csv".format(seed))
        train_df.astype({'Index': 'int32'}).to_csv(train_df_path, index=False)
        test_df_path = os.path.join(results_directory, "Testing_data seed_{}.csv".format(seed))
        test_df.astype({'Index': 'int32'}).to_csv(test_df_path, index=False)

        # Biasness Check
        svm_tn_h_indices, svm_fp_h_indices, svm_tp_d_indices, svm_fn_d_indices = svm.get_error_idx()
        mlp_tn_h_indices, mlp_fp_h_indices, mlp_tp_d_indices, mlp_fn_d_indices = mlp.get_error_idx()
        rf_tn_h_indices, rf_fp_h_indices, rf_tp_d_indices, rf_fn_d_indices = rf.get_error_idx()

        y_test_h = []
        y_test_d = []
        for idx, value in y_test.items():
            if value == 0:
                y_test_h.append(idx)
            elif value == 1:
                y_test_d.append(idx)

        if classification_mode == 'Standard':
            biasness_data = data
        elif classification_mode == 'Testing_Remaining':
            biasness_data = test_data_updated

        y_test_h_metadata = BiasnessData(biasness_data, y_test_h)
        y_test_h_metadata.get_metadata()
        y_test_h_age, y_test_h_genders, y_test_h_mmse, y_test_h_ad, y_test_h_dep = y_test_h_metadata.get_counts()
        y_test_d_metadata = BiasnessData(biasness_data, y_test_d)
        y_test_d_metadata.get_metadata()
        y_test_d_age, y_test_d_genders, y_test_d_mmse, y_test_d_ad, y_test_d_dep = y_test_d_metadata.get_counts()
        ytest_counter.update_counts(y_test_h_age, y_test_h_genders, y_test_h_mmse, y_test_h_ad, y_test_h_dep,
                                    y_test_d_age, y_test_d_genders, y_test_d_mmse, y_test_d_ad, y_test_d_dep)

        svm_tn_h_metadata = BiasnessData(biasness_data, svm_tn_h_indices)
        svm_tn_h_metadata.get_metadata()
        svm_tn_age, svm_tn_genders, svm_tn_mmse, svm_tn_ad, svm_tn_dep = svm_tn_h_metadata.get_counts()
        svm_fp_h_metadata = BiasnessData(biasness_data, svm_fp_h_indices)
        svm_fp_h_metadata.get_metadata()
        svm_fp_age, svm_fp_genders, svm_fp_mmse, svm_fp_ad, svm_fp_dep = svm_fp_h_metadata.get_counts()
        svm_tp_d_metadata = BiasnessData(biasness_data, svm_tp_d_indices)
        svm_tp_d_metadata.get_metadata()
        svm_tp_age, svm_tp_genders, svm_tp_mmse, svm_tp_ad, svm_tp_dep = svm_tp_d_metadata.get_counts()
        svm_fn_d_metadata = BiasnessData(biasness_data, svm_fn_d_indices)
        svm_fn_d_metadata.get_metadata()
        svm_fn_age, svm_fn_genders, svm_fn_mmse, svm_fn_ad, svm_fn_dep = svm_fn_d_metadata.get_counts()
        svm_counter.update_counts(svm_tn_age, svm_tn_genders, svm_tn_mmse, svm_tn_ad, svm_tn_dep,
                                  svm_fp_age, svm_fp_genders, svm_fp_mmse, svm_fp_ad, svm_fp_dep,
                                  svm_tp_age, svm_tp_genders, svm_tp_mmse, svm_tp_ad, svm_tp_dep,
                                  svm_fn_age, svm_fn_genders, svm_fn_mmse, svm_fn_ad, svm_fn_dep)

        mlp_tn_h_metadata = BiasnessData(biasness_data, mlp_tn_h_indices)
        mlp_tn_h_metadata.get_metadata()
        mlp_tn_age, mlp_tn_genders, mlp_tn_mmse, mlp_tn_ad, mlp_tn_dep = mlp_tn_h_metadata.get_counts()
        mlp_fp_h_metadata = BiasnessData(biasness_data, mlp_fp_h_indices)
        mlp_fp_h_metadata.get_metadata()
        mlp_fp_age, mlp_fp_genders, mlp_fp_mmse, mlp_fp_ad, mlp_fp_dep = mlp_fp_h_metadata.get_counts()
        mlp_tp_d_metadata = BiasnessData(biasness_data, mlp_tp_d_indices)
        mlp_tp_d_metadata.get_metadata()
        mlp_tp_age, mlp_tp_genders, mlp_tp_mmse, mlp_tp_ad, mlp_tp_dep = mlp_tp_d_metadata.get_counts()
        mlp_fn_d_metadata = BiasnessData(biasness_data, mlp_fn_d_indices)
        mlp_fn_d_metadata.get_metadata()
        mlp_fn_age, mlp_fn_genders, mlp_fn_mmse, mlp_fn_ad, mlp_fn_dep = mlp_fn_d_metadata.get_counts()
        mlp_counter.update_counts(mlp_tn_age, mlp_tn_genders, mlp_tn_mmse, mlp_tn_ad, mlp_tn_dep,
                                  mlp_fp_age, mlp_fp_genders, mlp_fp_mmse, mlp_fp_ad, mlp_fp_dep,
                                  mlp_tp_age, mlp_tp_genders, mlp_tp_mmse, mlp_tp_ad, mlp_tp_dep,
                                  mlp_fn_age, mlp_fn_genders, mlp_fn_mmse, mlp_fn_ad, mlp_fn_dep)

        rf_tn_h_metadata = BiasnessData(biasness_data, rf_tn_h_indices)
        rf_tn_h_metadata.get_metadata()
        rf_tn_age, rf_tn_genders, rf_tn_mmse, rf_tn_ad, rf_tn_dep = rf_tn_h_metadata.get_counts()
        rf_fp_h_metadata = BiasnessData(biasness_data, rf_fp_h_indices)
        rf_fp_h_metadata.get_metadata()
        rf_fp_age, rf_fp_genders, rf_fp_mmse, rf_fp_ad, rf_fp_dep = rf_fp_h_metadata.get_counts()
        rf_tp_d_metadata = BiasnessData(biasness_data, rf_tp_d_indices)
        rf_tp_d_metadata.get_metadata()
        rf_tp_age, rf_tp_genders, rf_tp_mmse, rf_tp_ad, rf_tp_dep = rf_tp_d_metadata.get_counts()
        rf_fn_d_metadata = BiasnessData(biasness_data, rf_fn_d_indices)
        rf_fn_d_metadata.get_metadata()
        rf_fn_age, rf_fn_genders, rf_fn_mmse, rf_fn_ad, rf_fn_dep = rf_fn_d_metadata.get_counts()
        rf_counter.update_counts(rf_tn_age, rf_tn_genders, rf_tn_mmse, rf_tn_ad, rf_tn_dep,
                                  rf_fp_age, rf_fp_genders, rf_fp_mmse, rf_fp_ad, rf_fp_dep,
                                  rf_tp_age, rf_tp_genders, rf_tp_mmse, rf_tp_ad, rf_tp_dep,
                                  rf_fn_age, rf_fn_genders, rf_fn_mmse, rf_fn_ad, rf_fn_dep)

    svm_tn_age_total, svm_tn_genders_total, svm_tn_mmse_total, svm_tn_ad_total, svm_tn_dep_total, \
    svm_fp_age_total, svm_fp_genders_total, svm_fp_mmse_total, svm_fp_ad_total, svm_fp_dep_total, \
    svm_tp_age_total, svm_tp_genders_total, svm_tp_mmse_total, svm_tp_ad_total, svm_tp_dep_total, \
    svm_fn_age_total, svm_fn_genders_total, svm_fn_mmse_total, svm_fn_ad_total, svm_fn_dep_total = svm_counter.get_total_counts()

    mlp_tn_age_total, mlp_tn_genders_total, mlp_tn_mmse_total, mlp_tn_ad_total, mlp_tn_dep_total, \
    mlp_fp_age_total, mlp_fp_genders_total, mlp_fp_mmse_total, mlp_fp_ad_total, mlp_fp_dep_total, \
    mlp_tp_age_total, mlp_tp_genders_total, mlp_tp_mmse_total, mlp_tp_ad_total, mlp_tp_dep_total, \
    mlp_fn_age_total, mlp_fn_genders_total, mlp_fn_mmse_total, mlp_fn_ad_total, mlp_fn_dep_total = mlp_counter.get_total_counts()

    rf_tn_age_total, rf_tn_genders_total, rf_tn_mmse_total, rf_tn_ad_total, rf_tn_dep_total, \
    rf_fp_age_total, rf_fp_genders_total, rf_fp_mmse_total, rf_fp_ad_total, rf_fp_dep_total, \
    rf_tp_age_total, rf_tp_genders_total, rf_tp_mmse_total, rf_tp_ad_total, rf_tp_dep_total, \
    rf_fn_age_total, rf_fn_genders_total, rf_fn_mmse_total, rf_fn_ad_total, rf_fn_dep_total = rf_counter.get_total_counts()

    y_test_h_age_total, y_test_h_genders_total, y_test_h_mmse_total, y_test_h_ad_total, y_test_h_dep_total, \
    y_test_d_age_total, y_test_d_genders_total, y_test_d_mmse_total, y_test_d_ad_total, y_test_d_dep_total = ytest_counter.get_total_counts()

    #-------- Biasness values relative to test set for for different classifiers ----------

    svm_tn_h_obj = BiasnessTable(svm_tn_age_total, svm_tn_genders_total, svm_tn_mmse_total, svm_tn_ad_total, svm_tn_dep_total,
                                  y_test_h_age_total, y_test_h_genders_total, y_test_h_mmse_total, y_test_h_ad_total, y_test_h_dep_total)
    svm_cc_h_age, svm_cc_h_gender, svm_cc_h_mmse, svm_cc_h_ad, svm_cc_h_dep = svm_tn_h_obj.get_table_data()
    svm_fp_h_obj = BiasnessTable(svm_fp_age_total, svm_fp_genders_total, svm_fp_mmse_total, svm_fp_ad_total, svm_fp_dep_total,
                                  y_test_h_age_total, y_test_h_genders_total, y_test_h_mmse_total, y_test_h_ad_total, y_test_h_dep_total)
    svm_mc_h_age, svm_mc_h_gender, svm_mc_h_mmse, svm_mc_h_ad, svm_mc_h_dep = svm_fp_h_obj.get_table_data()
    svm_tp_d_obj = BiasnessTable(svm_tp_age_total, svm_tp_genders_total, svm_tp_mmse_total, svm_tp_ad_total, svm_tp_dep_total,
                                  y_test_d_age_total, y_test_d_genders_total, y_test_d_mmse_total, y_test_d_ad_total, y_test_d_dep_total)
    svm_cc_d_age, svm_cc_d_gender, svm_cc_d_mmse, svm_cc_d_ad, svm_cc_d_dep = svm_tp_d_obj.get_table_data()
    svm_fn_d_obj = BiasnessTable(svm_fn_age_total, svm_fn_genders_total, svm_fn_mmse_total, svm_fn_ad_total, svm_fn_dep_total,
                                  y_test_d_age_total, y_test_d_genders_total, y_test_d_mmse_total, y_test_d_ad_total, y_test_d_dep_total)
    svm_mc_d_age, svm_mc_d_gender, svm_mc_d_mmse, svm_mc_d_ad, svm_mc_d_dep = svm_fn_d_obj.get_table_data()

    mlp_tn_h_obj = BiasnessTable(mlp_tn_age_total, mlp_tn_genders_total, mlp_tn_mmse_total, mlp_tn_ad_total, mlp_tn_dep_total,
                                  y_test_h_age_total, y_test_h_genders_total, y_test_h_mmse_total, y_test_h_ad_total, y_test_h_dep_total)
    mlp_cc_h_age, mlp_cc_h_gender, mlp_cc_h_mmse, mlp_cc_h_ad, mlp_cc_h_dep = mlp_tn_h_obj.get_table_data()
    mlp_fp_h_obj = BiasnessTable(mlp_fp_age_total, mlp_fp_genders_total, mlp_fp_mmse_total, mlp_fp_ad_total, mlp_fp_dep_total,
                                  y_test_h_age_total, y_test_h_genders_total, y_test_h_mmse_total, y_test_h_ad_total, y_test_h_dep_total)
    mlp_mc_h_age, mlp_mc_h_gender, mlp_mc_h_mmse, mlp_mc_h_ad, mlp_mc_h_dep = mlp_fp_h_obj.get_table_data()
    mlp_tp_d_obj = BiasnessTable(mlp_tp_age_total, mlp_tp_genders_total, mlp_tp_mmse_total, mlp_tp_ad_total, mlp_tp_dep_total,
                                  y_test_d_age_total, y_test_d_genders_total, y_test_d_mmse_total, y_test_d_ad_total, y_test_d_dep_total)
    mlp_cc_d_age, mlp_cc_d_gender, mlp_cc_d_mmse, mlp_cc_d_ad, mlp_cc_d_dep = mlp_tp_d_obj.get_table_data()
    mlp_fn_d_obj = BiasnessTable(mlp_fn_age_total, mlp_fn_genders_total, mlp_fn_mmse_total, mlp_fn_ad_total, mlp_fn_dep_total,
                                  y_test_d_age_total, y_test_d_genders_total, y_test_d_mmse_total, y_test_d_ad_total, y_test_d_dep_total)
    mlp_mc_d_age, mlp_mc_d_gender, mlp_mc_d_mmse, mlp_mc_d_ad, mlp_mc_d_dep = mlp_fn_d_obj.get_table_data()

    rf_tn_h_obj = BiasnessTable(rf_tn_age_total, rf_tn_genders_total, rf_tn_mmse_total, rf_tn_ad_total, rf_tn_dep_total,
                                  y_test_h_age_total, y_test_h_genders_total, y_test_h_mmse_total, y_test_h_ad_total, y_test_h_dep_total)
    rf_cc_h_age, rf_cc_h_gender, rf_cc_h_mmse, rf_cc_h_ad, rf_cc_h_dep = rf_tn_h_obj.get_table_data()
    rf_fp_h_obj = BiasnessTable(rf_fp_age_total, rf_fp_genders_total, rf_fp_mmse_total, rf_fp_ad_total, rf_fp_dep_total,
                                  y_test_h_age_total, y_test_h_genders_total, y_test_h_mmse_total, y_test_h_ad_total, y_test_h_dep_total)
    rf_mc_h_age, rf_mc_h_gender, rf_mc_h_mmse, rf_mc_h_ad, rf_mc_h_dep = rf_fp_h_obj.get_table_data()
    rf_tp_d_obj = BiasnessTable(rf_tp_age_total, rf_tp_genders_total, rf_tp_mmse_total, rf_tp_ad_total, rf_tp_dep_total,
                                  y_test_d_age_total, y_test_d_genders_total, y_test_d_mmse_total, y_test_d_ad_total, y_test_d_dep_total)
    rf_cc_d_age, rf_cc_d_gender, rf_cc_d_mmse, rf_cc_d_ad, rf_cc_d_dep = rf_tp_d_obj.get_table_data()
    rf_fn_d_obj = BiasnessTable(rf_fn_age_total, rf_fn_genders_total, rf_fn_mmse_total, rf_fn_ad_total, rf_fn_dep_total,
                                  y_test_d_age_total, y_test_d_genders_total, y_test_d_mmse_total, y_test_d_ad_total, y_test_d_dep_total)
    rf_mc_d_age, rf_mc_d_gender, rf_mc_d_mmse, rf_mc_d_ad, rf_mc_d_dep = rf_fn_d_obj.get_table_data()

    # ---------- Generate Results Table -----------------
    headers = ['Task', 'Age group 1', 'Age group 2', 'Males', 'Females', 'MMSE Low', 'MMSE Mild', 'MMSE Severe', 'Healthy Controls', 'AD Patients', 'Non-depressed', 'Depressed']

    task_list = ['SVM correctly classified HC', 'MLP correctly classified HC', 'RF correctly classified HC',
                 'SVM misclassified HC', 'MLP misclassified HC', 'RF misclassified HC',
                 'SVM correctly classified AD Patients', 'MLP correctly classified AD Patients', 'RF correctly classified AD Patients',
                 'SVM misclassified AD Patients', 'MLP misclassified AD Patients', 'RF misclassified AD Patients']

    age_group_1_list = [svm_cc_h_age[0], mlp_cc_h_age[0], rf_cc_h_age[0],
                        svm_mc_h_age[0], mlp_mc_h_age[0], rf_mc_h_age[0],
                        svm_cc_d_age[0], mlp_cc_d_age[0], rf_cc_d_age[0],
                        svm_mc_d_age[0], mlp_mc_d_age[0], rf_mc_d_age[0]]

    age_group_2_list = [svm_cc_h_age[1], mlp_cc_h_age[1], rf_cc_h_age[1],
                        svm_mc_h_age[1], mlp_mc_h_age[1], rf_mc_h_age[1],
                        svm_cc_d_age[1], mlp_cc_d_age[1], rf_cc_d_age[1],
                        svm_mc_d_age[1], mlp_mc_d_age[1], rf_mc_d_age[1]]

    males_list = [svm_cc_h_gender[0], mlp_cc_h_gender[0], rf_cc_h_gender[0],
                  svm_mc_h_gender[0], mlp_mc_h_gender[0], rf_mc_h_gender[0],
                  svm_cc_d_gender[0], mlp_cc_d_gender[0], rf_cc_d_gender[0],
                  svm_mc_d_gender[0], mlp_mc_d_gender[0], rf_mc_d_gender[0]]

    females_list = [svm_cc_h_gender[1], mlp_cc_h_gender[1], rf_cc_h_gender[1],
                    svm_mc_h_gender[1], mlp_mc_h_gender[1], rf_mc_h_gender[1],
                    svm_cc_d_gender[1], mlp_cc_d_gender[1], rf_cc_d_gender[1],
                    svm_mc_d_gender[1], mlp_mc_d_gender[1], rf_mc_d_gender[1]]

    mmse_low_list = [svm_cc_h_mmse[0], mlp_cc_h_mmse[0], rf_cc_h_mmse[0],
                     svm_mc_h_mmse[0], mlp_mc_h_mmse[0], rf_mc_h_mmse[0],
                     svm_cc_d_mmse[0], mlp_cc_d_mmse[0], rf_cc_d_mmse[0],
                     svm_mc_d_mmse[0], mlp_mc_d_mmse[0], rf_mc_d_mmse[0]]

    mmse_mild_list = [svm_cc_h_mmse[1], mlp_cc_h_mmse[1], rf_cc_h_mmse[1],
                      svm_mc_h_mmse[1], mlp_mc_h_mmse[1], rf_mc_h_mmse[1],
                      svm_cc_d_mmse[1], mlp_cc_d_mmse[1], rf_cc_d_mmse[1],
                      svm_mc_d_mmse[1], mlp_mc_d_mmse[1], rf_mc_d_mmse[1]]

    mmse_severe_list = [svm_cc_h_mmse[2], mlp_cc_h_mmse[2], rf_cc_h_mmse[2],
                        svm_mc_h_mmse[2], mlp_mc_h_mmse[2], rf_mc_h_mmse[2],
                        svm_cc_d_mmse[2], mlp_cc_d_mmse[2], rf_cc_d_mmse[2],
                        svm_mc_d_mmse[2], mlp_mc_d_mmse[2], rf_mc_d_mmse[2]]

    hc_list = [svm_cc_h_ad[0], mlp_cc_h_ad[0], rf_cc_h_ad[0],
               svm_mc_h_ad[0], mlp_mc_h_ad[0], rf_mc_h_ad[0],
               svm_cc_d_ad[0], mlp_cc_d_ad[0], rf_cc_d_ad[0],
               svm_mc_d_ad[0], mlp_mc_d_ad[0], rf_mc_d_ad[0]]

    ad_list = [svm_cc_h_ad[1], mlp_cc_h_ad[1], rf_cc_h_ad[1],
               svm_mc_h_ad[1], mlp_mc_h_ad[1], rf_mc_h_ad[1],
               svm_cc_d_ad[1], mlp_cc_d_ad[1], rf_cc_d_ad[1],
               svm_mc_d_ad[1], mlp_mc_d_ad[1], rf_mc_d_ad[1]]

    non_dep_list = [svm_cc_h_dep[0], mlp_cc_h_dep[0], rf_cc_h_dep[0],
                    svm_mc_h_dep[0], mlp_mc_h_dep[0], rf_mc_h_dep[0],
                    svm_cc_d_dep[0], mlp_cc_d_dep[0], rf_cc_d_dep[0],
                    svm_mc_d_dep[0], mlp_mc_d_dep[0], rf_mc_d_dep[0]]

    dep_list = [svm_cc_h_dep[1], mlp_cc_h_dep[1], rf_cc_h_dep[1],
                svm_mc_h_dep[1], mlp_mc_h_dep[1], rf_mc_h_dep[1],
                svm_cc_d_dep[1], mlp_cc_d_dep[1], rf_cc_d_dep[1],
                svm_mc_d_dep[1], mlp_mc_d_dep[1], rf_mc_d_dep[1]]

    df = pd.DataFrame(list(zip(task_list, age_group_1_list, age_group_2_list, males_list, females_list, mmse_low_list, mmse_mild_list, mmse_severe_list, hc_list, ad_list, non_dep_list, dep_list)),
                      columns=headers)

    df.to_excel(results_table_path, index=False)

    # Compute and save classifier metrics
    # Create DataFrames for each classifier's metrics
    svm_df = pd.DataFrame(svm_metrics_dict).T
    mlp_df = pd.DataFrame(mlp_metrics_dict).T
    rf_df = pd.DataFrame(rf_metrics_dict).T

    # Calculate summaries for each classifier
    svm_summary, svm_summary_combined = calculate_metrics_summary(svm_metrics_dict, random_seeds)
    mlp_summary, mlp_summary_combined = calculate_metrics_summary(mlp_metrics_dict, random_seeds)
    rf_summary, rf_summary_combined = calculate_metrics_summary(rf_metrics_dict, random_seeds)

    # create merged dataframes for each classifier
    svm_metrics_df = pd.concat([svm_df, svm_summary], axis=0)
    mlp_metrics_df = pd.concat([mlp_df, mlp_summary], axis=0)
    rf_metrics_df = pd.concat([rf_df, rf_summary], axis=0)

    # Write the DataFrames to the Excel file
    svm_metrics_df.to_excel(excel_writer, sheet_name='SVM', index=True, index_label='Random State')
    mlp_metrics_df.to_excel(excel_writer, sheet_name='MLP', index=True, index_label='Random State')
    rf_metrics_df.to_excel(excel_writer, sheet_name='RF', index=True, index_label='Random State')

    # Close the Pandas Excel writer and save the Excel file
    excel_writer.save()

    metrics_data = {
        "Embedding": [embedding_name, " ", " "],
        "Classifier": ["SVM", "MLP", "RF"],
        "Accuracy(%) ± std": [svm_summary_combined['accuracy'][0], mlp_summary_combined['accuracy'][0], rf_summary_combined['accuracy'][0]],
        "Unweighted average recall(%) ± std": [svm_summary_combined['uar'][0], mlp_summary_combined['uar'][0], rf_summary_combined['uar'][0]],
        "Sensitivity(%) ± std": [svm_summary_combined['sensitivity'][0], mlp_summary_combined['sensitivity'][0], rf_summary_combined['sensitivity'][0]],
        "Specificity(%) ± std": [svm_summary_combined['specificity'][0], mlp_summary_combined['specificity'][0], rf_summary_combined['specificity'][0]],
    }

    metrics_data_df = pd.DataFrame(metrics_data)
    print(metrics_data_df)

    book = load_workbook(existing_file)
    writer = pd.ExcelWriter(existing_file, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    metrics_data_df.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index=False, header=False)
    writer.save()


# --------------- main ------------------
if __name__ == "__main__":
    main()